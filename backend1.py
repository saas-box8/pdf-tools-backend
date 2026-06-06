"""
Excel → PDF Converter — Flask + Waitress

Run:
    pip install flask flask-cors openpyxl reportlab xlrd==1.2.0 waitress
    python app.py
"""

from __future__ import annotations

import os
import html
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import (
    Flowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Table,
    TableStyle,
)

try:
    import xlrd
except ImportError:
    xlrd = None

# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = Flask(__name__)
CORS(app, expose_headers=["Content-Disposition"])

HOST                = os.getenv("HOST", "0.0.0.0")
PORT                = int(os.getenv("PORT", "5000"))
THREADS             = int(os.getenv("THREADS", "8"))
MAX_CONTENT_MB      = int(os.getenv("MAX_CONTENT_MB", "25"))
MAX_REQUEST_BODY_MB = int(os.getenv("MAX_REQUEST_BODY_MB", "30"))
EXPOSE_TRACEBACKS   = os.getenv("EXPOSE_TRACEBACKS", "0").strip().lower() in {"1","true","yes","on"}

app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_MB * 1024 * 1024

# ---------------------------------------------------------------------------
# Quality settings
#
# `margin` is applied equally to ALL four sides (left = right = top = bottom).
# The footer lives inside the bottom margin so it never eats into content space.
# ---------------------------------------------------------------------------
FOOTER_HEIGHT = 0.22 * inch

QUALITY_SETTINGS = {
    "high": {
        "font_size":   10.5,
        "padding":     6,
        "line_width":  0.5,
        "margin":      0.18 * inch,
        "compression": 0,
    },
    "standard": {
        "font_size":   9.0,
        "padding":     4,
        "line_width":  0.35,
        "margin":      0.42 * inch,
        "compression": 1,
    },
    "compressed": {
        "font_size":   7.5,
        "padding":     2,
        "line_width":  0.2,
        "margin":      0.72 * inch,
        "compression": 1,
    },
}

SUPPORTED_FORMATS  = {"auto", "xls", "xlsx"}
OLE_XLS_SIGNATURE  = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
ZIP_SIGNATURE      = b"PK\x03\x04"


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------
def to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value)


def detect_excel_format(file_bytes, filename=""):
    head     = file_bytes[:8]
    filename = (filename or "").lower()
    if head.startswith(ZIP_SIGNATURE):      return "xlsx"
    if head.startswith(OLE_XLS_SIGNATURE):  return "xls"
    if filename.endswith(".xlsx"):          return "xlsx"
    if filename.endswith(".xls"):           return "xls"
    return None


def safe_download_name(original_filename: str) -> str:
    stem = Path(original_filename or "converted").stem.strip()
    return f"{stem or 'converted'}.pdf"


def trim_matrix(matrix):
    if not matrix:
        return []
    rows = []
    for row in matrix:
        row_text = [to_text(v).strip() for v in row]
        if any(cell != "" for cell in row_text):
            rows.append(row_text)
    if not rows:
        return []
    max_cols  = max(len(r) for r in rows)
    used_cols = [c for c in range(max_cols)
                 if any(c < len(r) and r[c].strip() != "" for r in rows)]
    if not used_cols:
        return []
    cleaned = []
    for r in rows:
        padded = r + [""] * (max_cols - len(r))
        cleaned.append([padded[c] for c in used_cols])
    return [r for r in cleaned if any(cell.strip() != "" for cell in r)]


def read_xlsx(file_bytes):
    wb     = load_workbook(BytesIO(file_bytes), data_only=True)
    sheets = []
    for ws in wb.worksheets:
        raw = [list(row) for row in ws.iter_rows(values_only=True)]
        sheets.append(trim_matrix(raw))
    return sheets


def read_xls(file_bytes):
    if xlrd is None:
        raise ValueError("xlrd is not installed. Run: pip install xlrd==1.2.0")
    book   = xlrd.open_workbook(file_contents=file_bytes)
    sheets = []
    for sh in book.sheets():
        raw = []
        for r in range(sh.nrows):
            row = []
            for c in range(sh.ncols):
                cell  = sh.cell(r, c)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(value, book.datemode)
                    except Exception:
                        pass
                row.append(value)
            raw.append(row)
        sheets.append(trim_matrix(raw))
    return sheets


def read_sheets_from_upload(file_storage, selected_format="auto"):
    filename   = file_storage.filename or ""
    file_bytes = file_storage.read()
    detected   = detect_excel_format(file_bytes, filename)

    if selected_format == "auto":
        selected_format = detected
    if selected_format not in SUPPORTED_FORMATS:
        raise ValueError("Unsupported Excel format selected.")
    if detected and selected_format and detected != selected_format:
        raise ValueError(
            f"File format mismatch. You selected {selected_format.upper()}, "
            f"but the file appears to be {detected.upper()}."
        )
    if selected_format == "xlsx": return read_xlsx(file_bytes)
    if selected_format == "xls":  return read_xls(file_bytes)
    raise ValueError("Could not detect the file format.")


# ---------------------------------------------------------------------------
# Column width — always fills exactly max_width (proportional distribution)
# ---------------------------------------------------------------------------
def estimate_col_widths(data, font_size, max_width):
    if not data:
        return [max_width]

    num_cols   = max((len(r) for r in data), default=1)
    raw_widths = []

    for col in range(num_cols):
        longest = 1
        for row in data[:80]:
            text    = to_text(row[col]) if col < len(row) else ""
            longest = max(longest, len(text))
        sample  = "W" * min(max(longest, 3), 40)
        approx  = stringWidth(sample, "Helvetica", font_size) * 0.60
        approx  = max(0.40 * inch, min(3.0 * inch, approx))
        raw_widths.append(approx)

    total = sum(raw_widths)
    scale = max_width / total
    return [w * scale for w in raw_widths]


def safe_paragraph_text(value):
    return html.escape(to_text(value)).replace("\n", "<br/>")


# ---------------------------------------------------------------------------
# Table builder — pure function, no page logic
# ---------------------------------------------------------------------------
def build_table(data, quality, gridlines, width):
    """Create a fully-styled ReportLab Table for *data* at the given *quality*.

    The table's column widths are scaled so it fills exactly *width*.
    Callers are responsible for deciding how to place the table in the
    document (single-page centered vs. multi-page split).
    """
    s          = QUALITY_SETTINGS[quality]
    font_size  = s["font_size"]
    padding    = s["padding"]
    line_width = s["line_width"]

    styles     = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "Cell",
        parent    = styles["BodyText"],
        fontName  = "Helvetica",
        fontSize  = font_size,
        leading   = font_size + 3,
        textColor = colors.black,
    )

    col_widths = estimate_col_widths(data, font_size, width)
    num_cols   = max((len(r) for r in data), default=1)
    if len(col_widths) != num_cols:
        col_widths = [width / max(num_cols, 1)] * num_cols

    table_data = []
    for row in data:
        rendered = []
        for i in range(num_cols):
            value = row[i] if i < len(row) else ""
            rendered.append(
                Paragraph(safe_paragraph_text(value) or "&nbsp;", cell_style)
            )
        table_data.append(rendered)

    # Repeat the header row on every page when the table splits
    repeat = 1 if len(table_data) > 1 else 0
    table  = Table(table_data, colWidths=col_widths, repeatRows=repeat)

    cmds = [
        ("BACKGROUND",    (0, 0), (-1,  0), colors.HexColor("#dbeafe")),
        ("FONTNAME",      (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 0), (-1, -1), font_size),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TEXTCOLOR",     (0, 0), (-1, -1), colors.black),
        ("LEFTPADDING",   (0, 0), (-1, -1), padding),
        ("RIGHTPADDING",  (0, 0), (-1, -1), padding),
        ("TOPPADDING",    (0, 0), (-1, -1), padding),
        ("BOTTOMPADDING", (0, 0), (-1, -1), padding),
    ]

    if gridlines:
        cmds.append(("GRID", (0, 0), (-1, -1), line_width, colors.HexColor("#94a3b8")))

    for r in range(1, len(table_data)):
        if r % 2 == 1:
            cmds.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#f8fafc")))

    if quality == "high":
        cmds.extend([
            ("LINEBELOW", (0, 0), (-1, 0), 1.1, colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ])

    table.setStyle(TableStyle(cmds))
    return table


# ---------------------------------------------------------------------------
# CenteredTable — single-page vertical centering wrapper
# ---------------------------------------------------------------------------
class CenteredTable(Flowable):
    """Wraps a pre-wrapped Table and draws it vertically centered inside
    the available frame.  Claims the full frame so nothing else shares
    the page."""

    def __init__(self, table, table_natural_height):
        super().__init__()
        self._table   = table
        self._table_h = table_natural_height
        self._y_off   = 0

    def wrap(self, availWidth, availHeight):
        # The offset that pushes the table into the vertical centre.
        # Clamped to 0 so an oversized table (shouldn't happen here)
        # simply top-aligns instead of going off-screen.
        self._y_off = max((availHeight - self._table_h) / 2.0, 0)
        return availWidth, availHeight          # occupy whole frame

    def draw(self):
        self.canv.saveState()
        self.canv.translate(0, self._y_off)
        self._table.drawOn(self.canv, 0, 0)
        self.canv.restoreState()


class BlankSheetPage(Flowable):
    """Placeholder for an empty sheet — produces a blank page."""
    def wrap(self, w, h): return w, h
    def draw(self):       pass


# ---------------------------------------------------------------------------
# PDF builder
# ---------------------------------------------------------------------------
def build_pdf(sheets, quality="standard", gridlines=True):
    if quality not in QUALITY_SETTINGS:
        quality = "standard"

    s          = QUALITY_SETTINGS[quality]
    margin     = s["margin"]
    page_size  = landscape(A4)
    page_w, page_h = page_size
    buffer     = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize        = page_size,
        leftMargin      = margin,
        rightMargin     = margin,
        topMargin       = margin,
        bottomMargin    = margin + FOOTER_HEIGHT,
        pageCompression = s["compression"],
    )

    # Exact frame dimensions the content area will have
    frame_w = page_w - margin * 2
    frame_h = page_h - margin * 2 - FOOTER_HEIGHT

    story = []
    for idx, data in enumerate(sheets):
        if idx > 0:
            story.append(PageBreak())

        if not data:
            story.append(BlankSheetPage())
            continue

        table = build_table(data, quality, gridlines, frame_w)

        # Wrap once to get the true natural height (respects quality
        # font-size, padding, leading, etc.)
        _, table_h = table.wrap(frame_w, frame_h)

        if table_h <= frame_h:
            # ── Fits on one page → center vertically ──
            story.append(CenteredTable(table, table_h))
        else:
            # ── Too tall → let ReportLab split it across pages ──
            #     (header row repeats thanks to repeatRows=1)
            story.append(table)

    if not story:
        raise ValueError("No visible data found.")

    footer_font = "Helvetica-Bold" if quality == "high" else "Helvetica"
    footer_size = 9                if quality == "high" else 8

    def add_footer(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont(footer_font, footer_size)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawCentredString(
            page_w / 2,
            FOOTER_HEIGHT * 0.45,
            f"Page {doc_obj.page}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------
@app.errorhandler(413)
def request_entity_too_large(_e):
    return jsonify({"error": f"File too large. Maximum allowed size is {MAX_CONTENT_MB} MB."}), 413

@app.errorhandler(404)
def not_found(_e):
    return jsonify({"error": "Not found."}), 404

@app.errorhandler(405)
def method_not_allowed(_e):
    return jsonify({"error": "Method not allowed."}), 405

@app.errorhandler(Exception)
def handle_unexpected_error(err):
    msg = str(err) if EXPOSE_TRACEBACKS else "Internal server error."
    return jsonify({"error": msg}), 500


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"ok": True, "message": "Excel to PDF server is running"}), 200


@app.route("/convert", methods=["POST"])
def convert():
    try:
        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"error": "No file uploaded."}), 400

        quality      = request.form.get("quality",      "standard").strip().lower()
        gridlines    = request.form.get("gridlines",    "yes"      ).strip().lower() == "yes"
        excel_format = request.form.get("excel_format", "auto"     ).strip().lower()

        if quality      not in QUALITY_SETTINGS: quality      = "standard"
        if excel_format not in SUPPORTED_FORMATS: excel_format = "auto"

        sheets     = read_sheets_from_upload(file, excel_format)
        pdf_buffer = build_pdf(sheets=sheets, quality=quality, gridlines=gridlines)
        pdf_name   = safe_download_name(file.filename)

        response = send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=pdf_name,
            mimetype="application/pdf",
            max_age=0,
        )
        response.headers["Cache-Control"] = "no-store"
        return response

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------
def main():
    try:
        from waitress import serve
    except ImportError as e:
        raise RuntimeError("pip install waitress") from e
    serve(
        app,
        host=HOST,
        port=PORT,
        threads=THREADS,
        max_request_body_size=MAX_REQUEST_BODY_MB * 1024 * 1024,
        channel_timeout=120,
        expose_tracebacks=EXPOSE_TRACEBACKS,
    )


if __name__ == "__main__":
    main()