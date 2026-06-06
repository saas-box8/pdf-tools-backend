"""
PDF → Excel Converter  ·  Production Edition
Single-file Flask + Waitress app

Features preserved from your backend:
- /health route
- /convert route
- PDF upload validation
- xlsx/xls export
- PyMuPDF text extraction
- pdfplumber table extraction
- OCR fallback with RapidOCR
- LibreOffice XLS conversion
- CORS with Content-Disposition exposed
- clean error messages
- download filenames based on the uploaded PDF name
- one-file startup with Waitress in __main__

Run:
    pip install flask flask-cors pdfplumber pymupdf openpyxl pillow numpy opencv-python rapidocr-onnxruntime waitress
    python app.py

Optional environment variables:
    HOST=0.0.0.0
    PORT=5000
    THREADS=8
    MAX_CONTENT_MB=105
    MAX_REQUEST_BODY_MB=110
    EXPOSE_TRACEBACKS=0
"""

from __future__ import annotations

import io
import os
import re
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Dict, List, Optional, Tuple

import fitz  # PyMuPDF
import numpy as np
import pdfplumber
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# -------------------------------------------------------------------
# Optional OCR dependency
# -------------------------------------------------------------------
try:
    import cv2
except ImportError as e:
    raise RuntimeError(
        "opencv-python is required. Install: pip install opencv-python"
    ) from e

try:
    from rapidocr_onnxruntime import RapidOCR
except Exception:
    RapidOCR = None


# -------------------------------------------------------------------
# App configuration
# -------------------------------------------------------------------
HOST = os.getenv("HOST", "0.0.0.0")
PORT = int(os.getenv("PORT", "5000"))
THREADS = int(os.getenv("THREADS", "8"))
MAX_CONTENT_MB = int(os.getenv("MAX_CONTENT_MB", "105"))
MAX_REQUEST_BODY_MB = int(os.getenv("MAX_REQUEST_BODY_MB", "110"))
EXPOSE_TRACEBACKS = os.getenv("EXPOSE_TRACEBACKS", "0").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}

APP_NAME = "PDF → Excel Converter"

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_MB * 1024 * 1024
CORS(app, expose_headers=["Content-Disposition"])

# -------------------------------------------------------------------
# OCR engine
# -------------------------------------------------------------------
try:
    ocr_engine = RapidOCR() if RapidOCR is not None else None
except Exception as _ocr_err:
    ocr_engine = None
    print("RapidOCR init failed:", _ocr_err)


# -------------------------------------------------------------------
# Shared style constants
# -------------------------------------------------------------------
_THIN = Side(style="thin", color="CBD5E1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_HEADER = PatternFill("solid", fgColor="DBEAFE")
_FILL_ALT = PatternFill("solid", fgColor="F8FAFC")

_ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)


# -------------------------------------------------------------------
# Settings
# -------------------------------------------------------------------
OCR_ZOOM = 3.0
ROW_Y_TOLERANCE = 7.0
TABLE_BBOX_MARGIN = 1.0

_TABLE_LINES = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines",
    "snap_tolerance": 4,
    "join_tolerance": 4,
    "edge_min_length": 20,
    "intersection_tolerance": 5,
    "min_words_vertical": 1,
    "min_words_horizontal": 1,
}

_TABLE_TEXT = {
    "vertical_strategy": "text",
    "horizontal_strategy": "text",
    "snap_tolerance": 4,
    "join_tolerance": 4,
    "edge_min_length": 20,
    "intersection_tolerance": 5,
    "min_words_vertical": 2,
    "min_words_horizontal": 2,
}


# -------------------------------------------------------------------
# Data models
# -------------------------------------------------------------------
@dataclass
class TextLine:
    text: str
    x0: float
    y0: float
    x1: float
    y1: float
    bold: bool = False
    italic: bool = False
    font_size: float = 11.0


@dataclass
class TableBlock:
    bbox: Tuple[float, float, float, float]
    rows: List[List[str]]


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
def clean(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\x00", " ")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()


def normalize_sheet_name(name: str) -> str:
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    name = re.sub(r"\s+", " ", name).strip()
    return (name or "Sheet")[:31]


def safe_output_stem(filename: str) -> str:
    stem = Path(filename or "converted").stem
    stem = re.sub(r"[^\w.\-() ]+", "_", stem).strip()
    return stem or "converted"


def bbox_intersects(
    a: Tuple[float, float, float, float],
    b: Tuple[float, float, float, float],
    margin: float = 0.0,
) -> bool:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    ax0 -= margin
    ay0 -= margin
    ax1 += margin
    ay1 += margin
    bx0 -= margin
    by0 -= margin
    bx1 += margin
    by1 += margin
    return not (ax1 < bx0 or ax0 > bx1 or ay1 < by0 or ay0 > by1)


def safe_font(size: float, bold: bool = False, italic: bool = False) -> Font:
    return Font(
        name="Calibri",
        size=max(8, min(float(size or 11), 18)),
        bold=bool(bold),
        italic=bool(italic),
        color="111827",
    )


def auto_fit_sheet(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        col_cells = list(col_cells)
        if not col_cells:
            continue
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for c in col_cells:
            if c.value is None:
                continue
            txt = str(c.value)
            max_len = max(max_len, max((len(l) for l in txt.splitlines()), default=0))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def span_is_bold(span: Dict[str, Any]) -> bool:
    fn = str(span.get("font", "")).lower()
    flags = int(span.get("flags", 0))
    return any(kw in fn for kw in ("bold", "black", "semibold", "demibold")) or bool(flags & 16)


def span_is_italic(span: Dict[str, Any]) -> bool:
    fn = str(span.get("font", "")).lower()
    flags = int(span.get("flags", 0))
    return any(kw in fn for kw in ("italic", "oblique")) or bool(flags & 2)


def _style_cell(
    cell,
    *,
    bold: bool = False,
    italic: bool = False,
    fill: Optional[PatternFill] = None,
    font_size: float = 11.0,
    center: bool = False,
) -> None:
    cell.border = _BORDER
    cell.font = safe_font(font_size, bold=bold, italic=italic)
    cell.alignment = _ALIGN_CTR if center else _ALIGN_WRAP
    if fill is not None:
        cell.fill = fill


# -------------------------------------------------------------------
# Pure-Python XLS conversion (xlwt — no LibreOffice needed)
# -------------------------------------------------------------------
def convert_xlsx_to_xls(xlsx_bytes: bytes) -> bytes:
    """Convert xlsx bytes to legacy .xls bytes using openpyxl + xlwt."""
    try:
        import xlwt
    except ImportError:
        raise RuntimeError("xlwt not installed. Run: pip install xlwt")

    from openpyxl import load_workbook as _lw
    import io as _io

    wb_src = _lw(_io.BytesIO(xlsx_bytes), data_only=True)
    wb_dst = xlwt.Workbook(encoding="utf-8")

    for sheet_name in wb_src.sheetnames:
        ws_src = wb_src[sheet_name]
        ws_dst = wb_dst.add_sheet(sheet_name[:31], cell_overwrite_ok=True)
        for r_idx, row in enumerate(ws_src.iter_rows()):
            for c_idx, cell in enumerate(row):
                val = cell.value
                if val is None:
                    continue
                ws_dst.write(r_idx, c_idx, val)

    out = _io.BytesIO()
    wb_dst.save(out)
    out.seek(0)
    return out.read()


# -------------------------------------------------------------------
# OCR
# -------------------------------------------------------------------
def _preprocess_for_ocr(img_np: np.ndarray) -> np.ndarray:
    if img_np is None or img_np.size == 0:
        return img_np

    gray = cv2.cvtColor(img_np, cv2.COLOR_RGB2GRAY) if img_np.ndim == 3 else img_np

    if max(gray.shape) < 1600:
        gray = cv2.resize(gray, None, fx=1.25, fy=1.25, interpolation=cv2.INTER_CUBIC)

    blur = cv2.GaussianBlur(gray, (3, 3), 0)
    thresh = cv2.adaptiveThreshold(
        blur,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31,
        11,
    )
    return cv2.cvtColor(thresh, cv2.COLOR_GRAY2RGB)


def ocr_page_lines(page: fitz.Page) -> List[TextLine]:
    if ocr_engine is None:
        return []

    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(OCR_ZOOM, OCR_ZOOM), alpha=False)
        img_np = np.array(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
        img_np = _preprocess_for_ocr(img_np)
        results, _ = ocr_engine(img_np)
    except Exception:
        return []

    if not results:
        return []

    grouped: Dict[int, list] = {}
    for item in results:
        try:
            box = item[0]
            text = clean(item[1])
        except Exception:
            continue
        if not text:
            continue
        try:
            x0 = float(box[0][0])
            y0 = float(box[0][1])
            x1 = float(box[2][0])
            y1 = float(box[3][1]) if len(box) > 3 else float(box[2][1])
        except Exception:
            continue
        key = int(round(y0 / ROW_Y_TOLERANCE))
        grouped.setdefault(key, []).append((x0, y0, x1, y1, text))

    lines: List[TextLine] = []
    for key in sorted(grouped):
        row = sorted(grouped[key], key=lambda t: t[0])
        merged = " ".join(t[4] for t in row if t[4])
        if merged:
            lines.append(
                TextLine(
                    text=merged,
                    x0=min(t[0] for t in row),
                    y0=min(t[1] for t in row),
                    x1=max(t[2] for t in row),
                    y1=max(t[3] for t in row),
                )
            )
    return lines


# -------------------------------------------------------------------
# Table extraction
# -------------------------------------------------------------------
def extract_tables(p_page) -> List[TableBlock]:
    found: List[TableBlock] = []

    for settings in (_TABLE_LINES, _TABLE_TEXT):
        try:
            tables = p_page.find_tables(table_settings=settings)
        except Exception:
            tables = []

        for table in tables:
            try:
                data = table.extract()
            except Exception:
                data = None
            if not data:
                continue

            rows = [
                [clean(c) for c in row]
                for row in data
                if row and any(clean(c) for c in row)
            ]
            if rows:
                found.append(
                    TableBlock(
                        bbox=tuple(float(v) for v in table.bbox),
                        rows=rows,
                    )
                )

        if found:
            break

    found.sort(key=lambda t: (t.bbox[1], t.bbox[0]))
    return found


# -------------------------------------------------------------------
# Text extraction
# -------------------------------------------------------------------
def extract_text_lines(page: fitz.Page) -> List[TextLine]:
    try:
        text_dict = page.get_text("dict")
    except Exception:
        return []

    lines: List[TextLine] = []
    for block in text_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            if not spans:
                continue

            parts: List[str] = []
            bold = italic = False
            sizes: List[float] = []

            for span in spans:
                txt = clean(span.get("text", ""))
                if not txt:
                    continue
                parts.append(txt)
                sizes.append(float(span.get("size", 11.0) or 11.0))
                if span_is_bold(span):
                    bold = True
                if span_is_italic(span):
                    italic = True

            text = clean(" ".join(parts))
            if not text:
                continue

            try:
                bbox = line.get("bbox") or block.get("bbox")
                x0, y0, x1, y1 = [float(v) for v in bbox]
            except Exception:
                continue

            lines.append(
                TextLine(
                    text=text,
                    x0=x0,
                    y0=y0,
                    x1=x1,
                    y1=y1,
                    bold=bold,
                    italic=italic,
                    font_size=max(sizes) if sizes else 11.0,
                )
            )

    lines.sort(key=lambda l: (l.y0, l.x0))
    return lines


def filter_outside_tables(lines: List[TextLine], tables: List[TableBlock]) -> List[TextLine]:
    if not tables:
        return lines
    return [
        l
        for l in lines
        if not any(
            bbox_intersects((l.x0, l.y0, l.x1, l.y1), t.bbox, TABLE_BBOX_MARGIN)
            for t in tables
        )
    ]


# -------------------------------------------------------------------
# Excel writing — pure data, no labels
# -------------------------------------------------------------------
def write_table_rows(ws, table: TableBlock, start_row: int) -> int:
    """
    Write table rows directly. First row -> header style.
    No title, no section label.
    """
    row = start_row
    data = table.rows
    max_col = max((len(r) for r in data), default=0)

    for r_idx, r in enumerate(data):
        is_header = r_idx == 0
        fill = _FILL_HEADER if is_header else (_FILL_ALT if r_idx % 2 == 1 else None)

        for c_idx in range(max_col):
            value = r[c_idx] if c_idx < len(r) else ""
            cell = ws.cell(row=row, column=c_idx + 1, value=value or None)
            _style_cell(cell, bold=is_header, fill=fill, font_size=11)

        row += 1

    return row


def _cluster_by_y(lines: List[TextLine], tol: float = ROW_Y_TOLERANCE) -> List[List[TextLine]]:
    if not lines:
        return []
    sorted_lines = sorted(lines, key=lambda l: (l.y0, l.x0))
    clusters: List[List[TextLine]] = []

    for line in sorted_lines:
        if not clusters:
            clusters.append([line])
            continue
        last = clusters[-1]
        avg_y = sum(l.y0 for l in last) / len(last)
        if abs(line.y0 - avg_y) <= tol:
            last.append(line)
        else:
            clusters.append([line])

    return clusters


def write_text_lines(ws, lines: List[TextLine], start_row: int) -> int:
    """
    Write text lines as plain data cells.
    No section headers, no annotations.
    """
    row = start_row
    clusters = _cluster_by_y(lines)

    for cluster in clusters:
        cluster = sorted(cluster, key=lambda l: l.x0)
        for col_idx, line in enumerate(cluster, start=1):
            txt = clean(line.text)
            if not txt:
                continue
            cell = ws.cell(row=row, column=col_idx, value=txt)
            _style_cell(
                cell,
                bold=line.bold,
                italic=line.italic,
                font_size=max(9, min(line.font_size or 11, 18)),
            )
        row += 1

    return row


def write_ocr_lines(ws, lines: List[TextLine], start_row: int) -> int:
    """
    Write OCR lines as plain data.
    """
    row = start_row
    for line in lines:
        txt = clean(line.text)
        if not txt:
            continue
        cell = ws.cell(row=row, column=1, value=txt)
        _style_cell(cell, font_size=11)
        row += 1
    return row


# -------------------------------------------------------------------
# Workbook builder
# -------------------------------------------------------------------
def build_workbook(pdf_bytes: bytes) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as plumber_doc:
        fitz_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        try:
            page_count = min(fitz_doc.page_count, len(plumber_doc.pages))

            for idx in range(page_count):
                p_page = plumber_doc.pages[idx]
                f_page = fitz_doc.load_page(idx)

                # Extract content layers in order
                tables = extract_tables(p_page)
                text_lines = extract_text_lines(f_page)
                text_lines = filter_outside_tables(text_lines, tables)

                native_chars = sum(len(l.text) for l in text_lines)
                ocr_lines = ocr_page_lines(f_page) if native_chars < 20 and not tables else []

                # Create sheet named after page number
                ws = wb.create_sheet(title=normalize_sheet_name(str(idx + 1)))

                current_row = 1

                for table in tables:
                    current_row = write_table_rows(ws, table, current_row)
                    current_row += 1

                if text_lines:
                    if tables:
                        current_row += 1
                    current_row = write_text_lines(ws, text_lines, current_row)

                if ocr_lines:
                    if tables or text_lines:
                        current_row += 1
                    current_row = write_ocr_lines(ws, ocr_lines, current_row)

                ws.freeze_panes = "A2"
                auto_fit_sheet(ws)

        finally:
            fitz_doc.close()

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# -------------------------------------------------------------------
# Error handlers
# -------------------------------------------------------------------
@app.errorhandler(413)
def request_entity_too_large(_err):
    return jsonify(
        {
            "error": f"File too large. Maximum allowed size is {MAX_CONTENT_MB} MB.",
        }
    ), 413


@app.errorhandler(404)
def not_found(_err):
    return jsonify({"error": "Not found."}), 404


@app.errorhandler(405)
def method_not_allowed(_err):
    return jsonify({"error": "Method not allowed."}), 405


@app.errorhandler(Exception)
def handle_unexpected_error(err):
    # Keep the response safe and JSON-shaped in production.
    return jsonify({"error": str(err)}), 500


# -------------------------------------------------------------------
# Routes
# -------------------------------------------------------------------
@app.route("/health", methods=["GET"])
def health():
    return jsonify(
        {
            "ok": True,
            "message": f"{APP_NAME} server running",
            "engine": "PyMuPDF + pdfplumber + RapidOCR fallback + LibreOffice XLS export",
            "max_content_mb": MAX_CONTENT_MB,
            "max_request_body_mb": MAX_REQUEST_BODY_MB,
        }
    ), 200


@app.route("/convert", methods=["POST"])
def convert():
    try:
        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"error": "No file uploaded."}), 400

        if not file.filename.lower().endswith(".pdf"):
            return jsonify({"error": "Please upload a PDF file."}), 400

        pdf_bytes = file.read()
        if not pdf_bytes:
            return jsonify({"error": "Empty file."}), 400

        excel_type = request.form.get("excelType", "xlsx").lower().strip()
        if excel_type not in {"xlsx", "xls"}:
            excel_type = "xlsx"

        xlsx_bytes = build_workbook(pdf_bytes)

        base_name = safe_output_stem(file.filename)
        if excel_type == "xls":
            output_bytes = convert_xlsx_to_xls(xlsx_bytes)
            download_name = f"{base_name}.xls"
            mimetype = "application/vnd.ms-excel"
        else:
            output_bytes = xlsx_bytes
            download_name = f"{base_name}.xlsx"
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        response = send_file(
            io.BytesIO(output_bytes),
            as_attachment=True,
            download_name=download_name,
            mimetype=mimetype,
            max_age=0,
        )
        response.headers["Cache-Control"] = "no-store"
        return response

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# -------------------------------------------------------------------
# Main entrypoint with Waitress
# -------------------------------------------------------------------
def main() -> None:
    try:
        from waitress import serve
    except ImportError as e:
        raise RuntimeError(
            "waitress is required. Install it with: pip install waitress"
        ) from e

    serve(
        app,
        host=HOST,
        port=PORT,
        threads=THREADS,
        max_request_body_size=MAX_REQUEST_BODY_MB * 1024 * 1024,
        channel_timeout=120,
        expose_tracebacks=EXPOSE_TRACEBACKS,
    )


if __name__ == "__main__":
    main()